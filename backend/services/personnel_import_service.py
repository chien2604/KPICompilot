"""Import Nghĩa Lâm personnel and Decision 283 KPI catalogs from XLSX files."""

import re
import shutil
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from ai_layer.rag.kuzu_graph_store import KuzuGraphStore
from core.config import Settings
from core.organization import (
    OUT_OF_SCOPE_ROLE,
    SPECIALIST_ROLE,
    UBND_AUTHORITY_ROLE,
    UNIT_DEPUTY_ROLE,
    UNIT_HEAD_ROLE,
    USER_ROLE,
    resolve_position_template,
)
from db.models.departments import Department
from db.models.kpi import WorkCatalogItem
from db.models.users import User, UserWorkArea
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from services.kpi_template_service import KPITemplateService

MISSING_INFORMATION = "Chưa cập nhật"
NON_PARTY_MEMBER = "Không ĐV"
EXPECTED_PERSONNEL_COUNT = 42
EXPECTED_WORK_ITEM_COUNT = 371

DEPARTMENT_NAMES = {
    "LANH_DAO_UBND": "Lãnh đạo UBND xã",
    "HDND_XA": "Hội đồng nhân dân xã",
    "VAN_PHONG_HDND_UBND": "Văn phòng HĐND - UBND",
    "PHONG_KINH_TE": "Phòng Kinh tế",
    "PHONG_VAN_HOA_XA_HOI": "Phòng Văn hóa - Xã hội",
    "TRUNG_TAM_HCC": "Trung tâm Phục vụ hành chính công",
    "TRUNG_TAM_CUDVC": "Trung tâm Cung ứng dịch vụ công",
}


@dataclass(frozen=True)
class PersonnelRecord:
    """Represent one normalized personnel row from the approved workbook."""

    department_code: str
    full_name: str
    date_of_birth: date | None
    ethnicity: str
    position_title: str
    source_work_area: str
    degree: str
    major: str
    political_theory: str
    personnel_type: str
    organization_role: str
    primary_position_code: str
    is_kpi_eligible: bool
    work_areas: tuple[tuple[str, str], ...]
    import_notes: str = ""


@dataclass(frozen=True)
class CommonCriterionRecord:
    """Represent one leaf criterion in Appendix I of Decision 283."""

    group_code: str
    group_name: str
    criterion_code: str
    criterion_name: str
    max_score: float
    sort_order: int


@dataclass(frozen=True)
class WorkCatalogRecord:
    """Represent one approved work product from Appendices II, III, and IV."""

    code: str
    catalog_scope: str
    department_code: str | None
    name: str
    details: str
    output: str
    complexity_group: str
    score_range: str
    conversion_score: float
    conversion_factor: float
    notes: str


class PersonnelWorkbookReader:
    """Read both personnel sheets without depending on row-specific names."""

    def read(self, workbook_path: Path) -> list[PersonnelRecord]:
        """Return validated personnel records from the UB and CUDVC sheets."""

        if not workbook_path.exists():
            raise FileNotFoundError(f"Không tìm thấy file nhân sự: {workbook_path}")
        workbook = load_workbook(workbook_path, data_only=True)
        required_sheets = {"UB", "Trung tâm CUDVC"}
        missing_sheets = required_sheets.difference(workbook.sheetnames)
        if missing_sheets:
            raise ValueError(f"File nhân sự thiếu sheet: {', '.join(missing_sheets)}")

        records = self._read_ub_sheet(workbook["UB"])
        records.extend(self._read_cudvc_sheet(workbook["Trung tâm CUDVC"]))
        if len(records) != EXPECTED_PERSONNEL_COUNT:
            raise ValueError(
                f"Danh sách phải có {EXPECTED_PERSONNEL_COUNT} người, đọc được {len(records)}."
            )
        names = [record.full_name for record in records]
        duplicate_names = sorted({name for name in names if names.count(name) > 1})
        if duplicate_names:
            raise ValueError(f"Họ tên bị trùng: {', '.join(duplicate_names)}")
        return records

    def _read_ub_sheet(self, worksheet) -> list[PersonnelRecord]:
        """Read elected officials, civil servants, and HCC staff from the UB sheet."""

        records: list[PersonnelRecord] = []
        for row in worksheet.iter_rows(min_row=9, values_only=True):
            if not isinstance(row[0], int) or not self._clean_text(row[1]):
                continue
            source_work_area = self._clean_text(row[7]) or MISSING_INFORMATION
            department_code = self._resolve_department_code(row[6], source_work_area)
            position_title = self._position_title(department_code, source_work_area)
            organization_role = self._organization_role(
                department_code, position_title
            )
            work_areas = self._work_areas(department_code, source_work_area)
            records.append(
                PersonnelRecord(
                    department_code=department_code,
                    full_name=self._clean_text(row[1]),
                    date_of_birth=self._date_value(row[2] or row[3]),
                    ethnicity=self._ethnicity(row[4], row[5]),
                    position_title=position_title,
                    source_work_area=source_work_area,
                    degree=self._clean_text(row[8]) or MISSING_INFORMATION,
                    major=self._clean_text(row[9]) or MISSING_INFORMATION,
                    political_theory=self._political_theory(row[10:13]),
                    personnel_type=(
                        "CAN_BO"
                        if department_code in {"LANH_DAO_UBND", "HDND_XA"}
                        else "CONG_CHUC"
                    ),
                    organization_role=organization_role,
                    primary_position_code=self._primary_position_code(
                        department_code, organization_role, source_work_area
                    ),
                    is_kpi_eligible=department_code not in {"LANH_DAO_UBND", "HDND_XA"},
                    work_areas=work_areas,
                    import_notes=self._source_ambiguity_note(row[4], row[5]),
                )
            )
        return records

    def _read_cudvc_sheet(self, worksheet) -> list[PersonnelRecord]:
        """Read five CUDVC staff and mark them outside the current KPI scope."""

        records: list[PersonnelRecord] = []
        for row in worksheet.iter_rows(min_row=10, values_only=True):
            if not isinstance(row[0], int) or not self._clean_text(row[1]):
                continue
            source_work_area = self._clean_text(row[21]) or MISSING_INFORMATION
            records.append(
                PersonnelRecord(
                    department_code="TRUNG_TAM_CUDVC",
                    full_name=self._clean_text(row[1]),
                    date_of_birth=self._date_value(row[2] or row[3]),
                    ethnicity=self._ethnicity(row[4], row[5]),
                    position_title=source_work_area,
                    source_work_area=source_work_area,
                    degree=self._clean_text(row[27]) or MISSING_INFORMATION,
                    major=self._clean_text(row[28]) or MISSING_INFORMATION,
                    political_theory=self._political_theory(row[29:32]),
                    personnel_type="VIEN_CHUC",
                    organization_role=OUT_OF_SCOPE_ROLE,
                    primary_position_code="CUDVC_CHUA_MATCH_TIEU_CHI",
                    is_kpi_eligible=False,
                    work_areas=(("CUDVC", source_work_area),),
                )
            )
        return records

    def _resolve_department_code(
        self, source_department: object, source_work_area: str
    ) -> str:
        """Map workbook department labels to stable internal codes."""

        value = self._normalize_search_text(source_department)
        if not value:
            title = self._normalize_search_text(source_work_area)
            return "LANH_DAO_UBND" if "ubnd" in title else "HDND_XA"
        if "vp hdnd ubnd" in value:
            return "VAN_PHONG_HDND_UBND"
        if value == "kinh te":
            return "PHONG_KINH_TE"
        if value == "vh xh":
            return "PHONG_VAN_HOA_XA_HOI"
        if "tt pv hcc" in value:
            return "TRUNG_TAM_HCC"
        raise ValueError(f"Không nhận diện được phòng/đơn vị: {source_department}")

    def _organization_role(self, department_code: str, position_title: str) -> str:
        """Derive assignment authority from explicit titles and unit membership."""

        if department_code == "LANH_DAO_UBND":
            return UBND_AUTHORITY_ROLE
        if department_code == "HDND_XA":
            return OUT_OF_SCOPE_ROLE
        normalized_title = self._normalize_search_text(position_title)
        if re.search(r"\bpho (truong|chanh)\b", normalized_title):
            return UNIT_DEPUTY_ROLE
        if re.search(r"\btruong (phong|don vi)\b", normalized_title):
            return UNIT_HEAD_ROLE
        if "chanh vp" in normalized_title:
            return UNIT_HEAD_ROLE
        if normalized_title.startswith("pgd") and "phu trach chung" in normalized_title:
            return UNIT_HEAD_ROLE
        return SPECIALIST_ROLE

    def _position_title(self, department_code: str, work_area: str) -> str:
        """Provide an explicit title when the source HCC cell is empty."""

        if work_area != MISSING_INFORMATION:
            return work_area
        if department_code == "TRUNG_TAM_HCC":
            return "Chuyên viên Trung tâm Phục vụ hành chính công"
        return "Chuyên viên"

    def _primary_position_code(
        self, department_code: str, organization_role: str, work_area: str
    ) -> str:
        """Create a stable primary-position code without encoding personnel names."""

        role_codes = {
            UBND_AUTHORITY_ROLE: "THAM_QUYEN_UBND",
            UNIT_HEAD_ROLE: "LDQL_TRUONG",
            UNIT_DEPUTY_ROLE: "LDQL_PHO",
            SPECIALIST_ROLE: "CMNV",
            OUT_OF_SCOPE_ROLE: "NGOAI_PHAM_VI",
        }
        area_code = self._work_areas(department_code, work_area)[0][0]
        return f"{role_codes[organization_role]}_{area_code}"

    def _work_areas(
        self, department_code: str, source_work_area: str
    ) -> tuple[tuple[str, str], ...]:
        """Match one primary position to one or more concurrent work areas."""

        normalized = self._normalize_search_text(source_work_area)
        if department_code == "LANH_DAO_UBND":
            return (("QL", "Lãnh đạo, quản lý"),)
        if department_code == "HDND_XA":
            return (("HDND", "Ngoài phạm vi KPI UBND"),)
        if department_code == "VAN_PHONG_HDND_UBND":
            return (("VP", "Văn phòng HĐND - UBND"),)
        if department_code == "TRUNG_TAM_HCC":
            return (("HCC", "Phục vụ hành chính công"),)
        if department_code == "TRUNG_TAM_CUDVC":
            return (("CUDVC", source_work_area),)

        areas: list[tuple[str, str]] = []
        if department_code == "PHONG_KINH_TE":
            keyword_mapping = (
                (("tai chinh", "ke toan", "ngan sach"), "KTTC", "Tài chính - kế hoạch"),
                (("xay dung", "quy hoach"), "KTXD", "Xây dựng - quy hoạch"),
                (("nong", "trong trot", "chan nuoi", "thuy san", "thuy loi"), "KTNN", "Nông nghiệp"),
                (("moi truong",), "KTMT", "Môi trường"),
                (("dat dai",), "KTDC", "Địa chính, đất đai"),
                (("giai phong mat bang", "gpmb"), "GPMB", "Giải phóng mặt bằng"),
                (("giam ngheo",), "ASXH", "Giảm nghèo, an sinh xã hội"),
            )
        else:
            keyword_mapping = (
                (("noi vu",), "NV", "Nội vụ"),
                (("van hoa", "thong tin"), "VHTT", "Văn hóa - thông tin"),
                (("giao duc",), "GD", "Giáo dục"),
                (("y te",), "YTE", "Y tế"),
                (("an sinh", "xa hoi"), "ASXH", "An sinh xã hội"),
            )
        for keywords, code, name in keyword_mapping:
            if any(keyword in normalized for keyword in keywords):
                areas.append((code, name))
        if not areas:
            fallback = "KT" if department_code == "PHONG_KINH_TE" else "VHXH"
            areas.append((fallback, source_work_area))
        return tuple(dict.fromkeys(areas))

    def _ethnicity(self, kinh_value: object, other_value: object) -> str:
        """Preserve all ethnicity marks and never write a null value."""

        values: list[str] = []
        if self._clean_text(kinh_value).lower() == "x":
            values.append("Kinh")
        other = self._clean_text(other_value)
        if other and other.lower() != "x":
            values.append(other)
        return ", ".join(values) or MISSING_INFORMATION

    def _source_ambiguity_note(self, kinh_value: object, other_value: object) -> str:
        """Flag contradictory ethnicity marks for later source verification."""

        if self._clean_text(kinh_value).lower() == "x" and self._clean_text(other_value):
            return "Nguồn đồng thời đánh dấu Kinh và dân tộc khác; cần đối chiếu hồ sơ gốc."
        return ""

    def _political_theory(self, values: tuple[object, ...]) -> str:
        """Convert marked theory columns into a Vietnamese level label."""

        labels = ("Sơ cấp", "Trung cấp", "Cao cấp")
        selected = [
            label
            for label, value in zip(labels, values, strict=False)
            if self._clean_text(value).lower() == "x"
        ]
        return ", ".join(selected) or MISSING_INFORMATION

    def _date_value(self, value: object) -> date | None:
        """Normalize Excel datetimes and dd/mm/yyyy text into a date."""

        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text_value = self._clean_text(value)
        if not text_value:
            return None
        return datetime.strptime(text_value, "%d/%m/%Y").date()

    def _normalize_search_text(self, value: object) -> str:
        """Normalize Vietnamese source text for stable keyword matching."""

        normalized = unicodedata.normalize("NFD", self._clean_text(value).lower())
        without_accents = "".join(
            character
            for character in normalized
            if unicodedata.category(character) != "Mn"
        )
        without_accents = without_accents.replace("đ", "d")
        return re.sub(r"[^a-z0-9]+", " ", without_accents).strip()

    def _clean_text(self, value: object) -> str:
        """Normalize Unicode and whitespace from one workbook cell."""

        if value is None:
            return ""
        return re.sub(r"\s+", " ", unicodedata.normalize("NFC", str(value))).strip()


class KPICatalogWorkbookReader:
    """Read common criteria and approved work products from Decision 283."""

    def read(
        self, workbook_path: Path
    ) -> tuple[list[CommonCriterionRecord], list[WorkCatalogRecord]]:
        """Return validated common criteria and the complete work catalog."""

        if not workbook_path.exists():
            raise FileNotFoundError(f"Không tìm thấy file tiêu chí: {workbook_path}")
        workbook = load_workbook(workbook_path, data_only=True)
        criteria = self._read_common_criteria(workbook["PL I - Tieu chi chung"])
        catalog: list[WorkCatalogRecord] = []
        catalog.extend(
            self._read_catalog_sheet(workbook["PL II - Lanh dao QL"], "LEADERSHIP")
        )
        catalog.extend(
            self._read_catalog_sheet(workbook["PL III - Dung chung"], "COMMON")
        )
        catalog.extend(
            self._read_catalog_sheet(workbook["PL IV - Rieng phong"], "DEPARTMENT")
        )
        if round(sum(item.max_score for item in criteria), 2) != 30:
            raise ValueError("Tổng điểm tiêu chí chung phải bằng 30.")
        if len(catalog) != EXPECTED_WORK_ITEM_COUNT:
            raise ValueError(
                f"Danh mục phải có {EXPECTED_WORK_ITEM_COUNT} mã, đọc được {len(catalog)}."
            )
        codes = [item.code for item in catalog]
        duplicate_codes = sorted({code for code in codes if codes.count(code) > 1})
        if duplicate_codes:
            raise ValueError(f"Mã công việc bị trùng: {', '.join(duplicate_codes)}")
        for item in catalog:
            if abs(item.conversion_factor - item.conversion_score / 5) > 0.0001:
                raise ValueError(f"Hệ số quy đổi không khớp tại mã {item.code}.")
        return criteria, catalog

    def _read_common_criteria(self, worksheet) -> list[CommonCriterionRecord]:
        """Read only leaf rows while retaining their parent criterion groups."""

        records: list[CommonCriterionRecord] = []
        group_code = ""
        group_name = ""
        for row in worksheet.iter_rows(min_row=4, values_only=True):
            sequence, name, score = row[:3]
            if isinstance(sequence, str) and re.fullmatch(r"[IVX]+", sequence):
                group_code = sequence
                group_name = str(name).strip()
                continue
            if isinstance(sequence, int) and name and score is not None:
                records.append(
                    CommonCriterionRecord(
                        group_code=group_code,
                        group_name=group_name,
                        criterion_code=f"TC.{group_code}.{sequence}",
                        criterion_name=str(name).strip(),
                        max_score=float(score),
                        sort_order=len(records) + 1,
                    )
                )
        return records

    def _read_catalog_sheet(
        self, worksheet, catalog_scope: str
    ) -> list[WorkCatalogRecord]:
        """Read catalog rows from one appendix using its stable code column."""

        records: list[WorkCatalogRecord] = []
        for row in worksheet.iter_rows(values_only=True):
            code = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not re.fullmatch(r"[A-Z]+(?:\.[A-Z]+)*\.\d+|[A-Z]+\.\d+", code):
                continue
            offset = 1 if catalog_scope == "DEPARTMENT" else 0
            source_department = str(row[2]).strip() if offset else ""
            name = str(row[2 + offset]).strip()
            details = str(row[3 + offset]).strip()
            output = str(row[4 + offset]).strip()
            complexity_group = str(row[5 + offset]).strip()
            score_range = str(row[6 + offset]).strip()
            conversion_score = float(row[7 + offset])
            conversion_factor = float(row[8 + offset])
            notes = str(row[9 + offset]).strip() if row[9 + offset] else ""
            records.append(
                WorkCatalogRecord(
                    code=code,
                    catalog_scope=catalog_scope,
                    department_code=self._catalog_department_code(source_department),
                    name=name,
                    details=details,
                    output=output,
                    complexity_group=complexity_group,
                    score_range=score_range,
                    conversion_score=conversion_score,
                    conversion_factor=conversion_factor,
                    notes=notes,
                )
            )
        return records

    def _catalog_department_code(self, source_department: str) -> str | None:
        """Map Appendix IV department labels to imported department codes."""

        mapping = {
            "Văn phòng HĐND-UBND": "VAN_PHONG_HDND_UBND",
            "Phòng Kinh tế": "PHONG_KINH_TE",
            "Phòng Văn hóa - Xã hội": "PHONG_VAN_HOA_XA_HOI",
            "Trung tâm Phục vụ hành chính công": "TRUNG_TAM_HCC",
        }
        if not source_department:
            return None
        if source_department not in mapping:
            raise ValueError(f"Phòng trong danh mục chưa được ánh xạ: {source_department}")
        return mapping[source_department]


class PersonnelImportService:
    """Atomically replace PoC business data with approved Nghĩa Lâm sources."""

    def __init__(self, database_session: Session, settings: Settings) -> None:
        """Store database, settings, and workbook reader dependencies."""

        self.database_session = database_session
        self.settings = settings
        self.personnel_reader = PersonnelWorkbookReader()
        self.catalog_reader = KPICatalogWorkbookReader()

    def reset_and_import(self) -> dict:
        """Validate both workbooks, reset business data, import, and synchronize graph."""

        personnel_records = self.personnel_reader.read(
            self.settings.personnel_import_path
        )
        criteria, catalog = self.catalog_reader.read(
            self.settings.work_catalog_import_path
        )
        self._truncate_business_tables()
        KPITemplateService(self.database_session).replace_templates(criteria)
        organization = self._create_organization()
        departments = self._create_departments(organization)
        self._create_work_catalog(catalog)
        users = self._create_users(departments, personnel_records)
        self._configure_reporting_lines(users)
        self.database_session.commit()
        self._clear_local_storage()
        self._synchronize_graph(organization, departments, users)
        return {
            "organization": organization.name,
            "department_count": len(departments),
            "personnel_count": len(users),
            "kpi_eligible_count": sum(user.is_kpi_eligible for user in users),
            "common_criterion_count": len(criteria),
            "work_catalog_count": len(catalog),
            "active_account_count": 0,
        }

    def _truncate_business_tables(self) -> None:
        """Remove replaceable business data while preserving the database schema."""

        self.database_session.execute(
            text(
                """
                TRUNCATE TABLE
                    conversation_summary, messages, conversations, chat_logs,
                    reports, kpi_scores, kpi_assessment_inputs, kpi_criteria,
                    document_type_rules, kpi_templates, document_chunks,
                    task_evidences, task_assignments, tasks, work_catalog_items,
                    user_work_areas, users, departments
                RESTART IDENTITY CASCADE
                """
            )
        )

    def _create_organization(self) -> Department:
        """Create the UBND xã Nghĩa Lâm organization root."""

        organization = Department(
            name=self.settings.organization_name,
            code=self.settings.organization_code,
            unit_type="ORGANIZATION",
            parent_id=None,
        )
        self.database_session.add(organization)
        self.database_session.flush()
        return organization

    def _create_departments(self, organization: Department) -> dict[str, Department]:
        """Create all personnel units in a deterministic display order."""

        departments: dict[str, Department] = {}
        for code, name in DEPARTMENT_NAMES.items():
            department = Department(
                name=name,
                code=code,
                unit_type=(
                    "AUTHORITY" if code == "LANH_DAO_UBND"
                    else "OUT_OF_SCOPE" if code == "HDND_XA"
                    else "UNIT"
                ),
                parent_id=organization.id,
            )
            self.database_session.add(department)
            self.database_session.flush()
            departments[code] = department
        return departments

    def _create_work_catalog(self, catalog: list[WorkCatalogRecord]) -> None:
        """Persist every approved product and its conversion factor."""

        for item in catalog:
            self.database_session.add(WorkCatalogItem(**item.__dict__))
        self.database_session.flush()

    def _create_users(
        self,
        departments: dict[str, Department],
        records: list[PersonnelRecord],
    ) -> list[User]:
        """Create inactive profiles and their normalized concurrent work areas."""

        users: list[User] = []
        for record in records:
            template = resolve_position_template(record.organization_role)
            user = User(
                full_name=record.full_name,
                email=None,
                hashed_password=None,
                role=USER_ROLE,
                kpi_role_template=template.code,
                permission_level=template.permission_level,
                organization_role=record.organization_role,
                organization_domain=(
                    "HDND" if record.department_code == "HDND_XA"
                    else "OUT_OF_UBND_KPI_SCOPE" if record.department_code == "TRUNG_TAM_CUDVC"
                    else "UBND"
                ),
                primary_position_code=record.primary_position_code,
                personnel_type=record.personnel_type,
                is_kpi_eligible=record.is_kpi_eligible,
                department_id=departments[record.department_code].id,
                position_title=record.position_title,
                phone_number=None,
                birth_year=record.date_of_birth.year if record.date_of_birth else None,
                date_of_birth=record.date_of_birth,
                ethnicity=record.ethnicity,
                party_joined_date=NON_PARTY_MEMBER,
                general_education=MISSING_INFORMATION,
                professional_qualification=f"{record.degree} - {record.major}",
                political_theory=record.political_theory,
                source_work_area=record.source_work_area,
                import_notes=record.import_notes,
                avatar_url=None,
                is_active=False,
            )
            self.database_session.add(user)
            self.database_session.flush()
            for area_index, (area_code, area_name) in enumerate(record.work_areas):
                self.database_session.add(
                    UserWorkArea(
                        user_id=user.id,
                        area_code=area_code,
                        area_name=area_name,
                        is_primary=area_index == 0,
                    )
                )
            users.append(user)
        self.database_session.flush()
        return users

    def _configure_reporting_lines(self, users: list[User]) -> None:
        """Link authority, unit heads, deputies, and specialists after bootstrap."""

        authority = next(
            (
                user
                for user in users
                if user.organization_role == UBND_AUTHORITY_ROLE
                and "phó" not in user.position_title.lower()
            ),
            None,
        )
        heads = {
            user.department_id: user
            for user in users
            if user.organization_role == UNIT_HEAD_ROLE
        }
        for user in users:
            if user.organization_role == UNIT_HEAD_ROLE:
                user.manager_id = authority.id if authority else None
                user.management_scope_json = {"all_department": True}
            elif user.organization_role in {UNIT_DEPUTY_ROLE, SPECIALIST_ROLE}:
                head = heads.get(user.department_id)
                user.manager_id = head.id if head else None
            if user.organization_role == UNIT_DEPUTY_ROLE:
                user.management_scope_json = {
                    "all_department": "phụ trách chung" in user.position_title.lower(),
                    "work_area_codes": [area.area_code for area in user.work_areas],
                }
        self.database_session.flush()

    def _clear_local_storage(self) -> None:
        """Remove old evidence files and rebuild the embedded graph directory."""

        for upload_path in self.settings.upload_dir.iterdir():
            if upload_path.name == ".gitkeep":
                continue
            if upload_path.is_dir():
                shutil.rmtree(upload_path)
            else:
                upload_path.unlink()
        if self.settings.kuzu_db_path.exists():
            shutil.rmtree(self.settings.kuzu_db_path)
        self.settings.kuzu_db_path.mkdir(parents=True, exist_ok=True)

    def _synchronize_graph(
        self,
        organization: Department,
        departments: dict[str, Department],
        users: list[User],
    ) -> None:
        """Synchronize organization and personnel nodes to embedded KùzuDB."""

        graph_store = KuzuGraphStore(str(self.settings.kuzu_db_path))
        graph_store.init_schema()
        graph_store.upsert_department(organization)
        for department in departments.values():
            graph_store.upsert_department(department)
        for user in users:
            graph_store.upsert_user(user)
            graph_store.link_user_department(user.id, user.department_id)
