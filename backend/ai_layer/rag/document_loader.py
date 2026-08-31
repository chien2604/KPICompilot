from pathlib import Path


class DocumentLoader:
    """Extract searchable text from supported evidence files."""

    def extract_text(self, file_path: str) -> str:
        """Extract text by file extension and remove PostgreSQL-incompatible null bytes."""

        path = Path(file_path)
        suffix = path.suffix.lower()
        try:
            if suffix in {".txt", ".md", ".csv"}:
                text = path.read_text(encoding="utf-8", errors="ignore")
            elif suffix == ".pdf":
                text = self._extract_pdf(path)
            elif suffix == ".docx":
                text = self._extract_docx(path)
            elif suffix in {".xlsx", ".xls"}:
                text = self._extract_excel(path)
            else:
                # Không đọc trực tiếp file nhị phân chưa hỗ trợ thành text
                text = f"Đã upload file {path.name} định dạng {suffix} chưa được hỗ trợ trích xuất toàn văn."
        except Exception as error:
            text = f"Không đọc được nội dung file {path.name}: {error}"

        # Bắt buộc: Loại bỏ ký tự null (\x00) vì PostgreSQL không chấp nhận ký tự này trong trường TEXT
        return text.replace("\x00", "")

    def _extract_excel(self, path: Path) -> str:
        """Extract cell values from XLS or XLSX workbooks."""

        if path.suffix.lower() == ".xls":
            import xlrd

            workbook = xlrd.open_workbook(str(path))
            text_parts: list[str] = []
            for worksheet in workbook.sheets():
                text_parts.append(f"--- Sheet: {worksheet.name} ---")
                for row_index in range(worksheet.nrows):
                    row_values = [
                        str(worksheet.cell_value(row_index, column_index))
                        for column_index in range(worksheet.ncols)
                    ]
                    text_parts.append("\t".join(row_values))
            return "\n".join(text_parts)

        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        text_parts = []
        for worksheet in workbook.worksheets:
            text_parts.append(f"--- Sheet: {worksheet.title} ---")
            for row in worksheet.iter_rows(values_only=True):
                text_parts.append(
                    "\t".join("" if value is None else str(value) for value in row)
                )
        workbook.close()
        return "\n".join(text_parts)

    def _extract_pdf(self, path: Path) -> str:
        """Extract PDF text with Docling and fall back to pypdf."""

        try:
            from docling.document_converter import DocumentConverter

            result = DocumentConverter().convert(str(path))
            return result.document.export_to_markdown()
        except Exception:
            try:
                from pypdf import PdfReader

                reader = PdfReader(str(path))
                return "\n".join(page.extract_text() or "" for page in reader.pages)
            except Exception:
                return (
                    f"PDF {path.name} đã được upload nhưng chưa trích xuất được text."
                )

    def _extract_docx(self, path: Path) -> str:
        """Extract DOCX text with Docling and fall back to python-docx."""

        try:
            from docling.document_converter import DocumentConverter

            result = DocumentConverter().convert(str(path))
            return result.document.export_to_markdown()
        except Exception:
            try:
                from docx import Document

                doc = Document(str(path))
                return "\n".join(p.text for p in doc.paragraphs)
            except Exception:
                return (
                    f"DOCX {path.name} đã được upload nhưng chưa trích xuất được text."
                )
